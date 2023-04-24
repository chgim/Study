from openpyxl import load_workbook

wb = load_workbook('result.xlsx') # 기존 워크북 불러오기
ws = wb.active # 마지막으로 열었던 Worksheet

print(ws['A1'].value) # A1에 있는 데이터 가져오기
print(ws['A:B']) # 슬라이싱을 통해 셀 범위에 접근할 수 있다.
print(ws['1:2']) 

################
for row in ws.iter_rows(values_only=True): # 한 행씩 가져오는 함수: iter_rows(…) values_only 파라미터를 취해 셀의 값만 반환할 수 있다.
    print(row)

# load_workbook: 모든 엑셀의 내용을 파이썬으로 한번에 가져옴
# ‣ 엑셀 파일이 매우 큰 경우 못 가져오는 경우 발생
# ‣ 한번에 가져오는 과정이 매우 느림