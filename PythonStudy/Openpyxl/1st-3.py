from openpyxl import Workbook
# 사용자가 접근해야 할 엑셀 파일을 설정하겠다는 의미입니다.
#  Workbook을 생성하는 경우는 2가지 케이스가 있습니다. 새로운 엑셀 파일을 만들거나, 원래 있는 파일에 접근하는 경우입니다.
wb = Workbook() # 데이터를 쓸 클래스 변수 생성 (파일명 없이)
ws = wb.active # 마지막으로 열었던 Worksheet
# ws = wb.create_sheet('sheet_test')    클래스 변수에서 시트이름으로 원하는 시트를 가져오거나 생성
ws['A1'] = 'Test Data!' # 원하는 위치에 데이터를 쓴후 저장한다.
wb.save('result.xlsx')


# python -m venv venv
# 가상 환경 만들기
# TERMINAL-> Command Prompt-> pip install openpyxl
# openpyxl 라이브러리 설치