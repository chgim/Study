import random
from openpyxl import Workbook, load_workbook #Workbook: 사용자가 접근해야 할 엑셀 파일을 설정, loac_workbook: 액셀 파일 가져오기

wb = Workbook()
ws = wb.active # 현재 Active Sheet 얻기

for _ in range(10000): # 10000 개의 행
    # row = []
    # for _ in range(10):
    #     row.append(random.randint(1, 100))
    row = [random.randint(1, 100) for _ in range(10)]  #row에 1~100 사이의 랜덤한 숫자를 10번 저장(10개의 열)
    ws.append(row) # ws에 row 추가

wb.save('result.xlsx') # ws를 result.xlst라는 파일에 저장

#---------------------------------------------------------------------------------------------------------#

wb = load_workbook('result.xlsx') # result.xlsx 파일에서 가져오기
ws = wb.active # 현재 Active Sheet 얻기

calc = [0,] * 10
for row in ws.iter_rows(values_only=True):
    for idx in range(10): # 10회 반복
        calc[idx] += row[idx] 

ws.append(calc) # ws에 calc 결과 추가
wb.save('result.xlsx')  # ws를 result.xlst라는 파일에 저장




# 1. 엑셀 파일에  다음 데이터를 생성

# - 10개의 열 10000개의 행

# - 각 셀에는 랜덤으로 1~100 사이의 숫자를 저장

# 2. result.xlsx 파일에 저장

#-----------------------------------------------------------#

# 3. load_workbook 으로 해당 파일을 다시 열기

# 4. 각 행의 값을 읽어서 10001 번째 행에 각 컬럼 별 합계를 추가

# 5. result.xlsx 파일에 저장



# result.xlsx파일과 python 파일을 압축하여 제출